# the file system so that we can crud excel files 

# dependencies
import openpyxl, os, sys

BASE_PATH = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) # adds project dir to places it looks for the modules
sys.path.append(BASE_PATH)

class Data():
    """the fs for excel files"""
    def __init__(self, filename, ws_name):
        self.filename = filename
        self.file_path = self.filename
        self.wb = openpyxl.load_workbook(self.file_path)
        self.ws = ""
        
        # set the active sheet to the one requested, if it does not exist: err
        
        sheets =  self.wb.sheetnames
        for sheet in sheets:
            if ws_name == sheet:
                self.ws = self.wb[sheet]

        if self.ws == "":
            raise Exception('That sheet name does not exist')


    def return_column_as_list(self, column_letter):
        """get a column by letter, each row is an element in a list"""
        # sanity check column_letter
        acceptable_letters = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
        if not column_letter in acceptable_letters: # note, this means AA is invalid
            raise Exception("Invalid column letter(s)")
        
        column = self.ws[column_letter]
        column_list = []
        for cell in column:
            val = cell.value if cell.value != None else ""
            column_list.append(val)
        return column_list

    def update_all_cells_in_column(self, column_letter, new_cell_values):
        """update a specific cell to a new value"""
        # sanity check column_letter
        acceptable_letters = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
        if not column_letter in acceptable_letters: # note, this means AA is invalid
            raise Exception("Invalid column letter(s)")

        # sanity check new_cell_values
        if type(new_cell_values) != list or len(new_cell_values) < 1:
            raise Exception("new_cell_values must be list that is not empty")

        # get column
        column = self.ws[column_letter]

        # make sure no new cell are being created (this is only update old cells)
        if  len(column) < len(new_cell_values):
            raise Exception("new_cell_values had more values than there were cells in column " + column_letter)

        # loop through each new cell value
        for x in range(len(new_cell_values)):
            column[x].value = new_cell_values[x]

        # save changes
        self.wb.save(self.file_path)

# TODO del me
"""
data = Data("Book1","poopPy")
print("---------------------------------------------------------")
col = data.return_column_as_list('A')
print(col)
print("---------------------------------------------------------")
print("---------------------------------------------------------")
data.update_all_cells_in_column('A',["joe","is","cool"])
col = data.return_column_as_list('A')
print(col)
"""